#!/usr/bin/env python3
"""
Gera a planilha de memória de cálculo do Estudo 11 (série histórica e
projeção do IBS total, Brasil, 2029-2033) a partir de
data/ibs-projecao-nacional.json e data/macro-parametros.json.

Saída: materiais/ibs-projecao-nacional-memoria-calculo.xlsx, com abas:
  - Síntese: pergunta, resposta e fontes
  - Série histórica 2013-2025: ICMS, ISS, arrecadação total, PIB, %PIB (dados de entrada)
  - Premissas macro 2026-2033: PIB real, IPCA, fonte (dados de entrada)
  - PIB projetado: fórmulas que compõem o PIB real ano a ano (R$ constantes de 2025)
  - Projeção IBS 2029-2033: fórmulas que aplicam o cronograma ADCT

Ao contrário das abas de dados de entrada (série histórica, premissas
macro), as abas "PIB projetado" e "Projeção IBS" usam fórmulas de planilha
que referenciam as células de entrada, não valores fixos. Mudar uma
premissa recalcula a projeção inteira.

Uso:
  python3 data/build-ibs-projecao-nacional-xlsx.py
"""

import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
PROJ = json.loads((HERE / "ibs-projecao-nacional.json").read_text())
MACRO = json.loads((HERE / "macro-parametros.json").read_text())
OUTPUT = HERE.parent / "materiais" / "ibs-projecao-nacional-memoria-calculo.xlsx"

NAVY = "1A3A5C"
LIGHT_BG = "EDF2F7"
WHITE = "FFFFFF"
INPUT_BLUE = "0000FF"

TITLE_FONT = Font(name="Calibri", size=15, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="6E6E6E")
H_FONT = Font(name="Calibri", size=11, bold=True, color="1D1D1B")
BODY_FONT = Font(name="Calibri", size=10.5, color="1D1D1B")
NOTE_FONT = Font(name="Calibri", size=9.5, italic=True, color="6E6E6E")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Calibri", size=10.5, bold=True, color=WHITE)
TOTAL_FILL = PatternFill("solid", fgColor=LIGHT_BG)
TOTAL_FONT = Font(name="Calibri", size=10.5, bold=True)
INPUT_FONT = Font(name="Calibri", size=10.5, color=INPUT_BLUE)
FORMULA_FONT = Font(name="Calibri", size=10.5, color="000000")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RS_FMT = '"R$" #,##0.0'
RS_FMT_FULL = '"R$" #,##0'
PCT_FMT = '0.00%'
PCT_FMT_SIGNED = '+0.00%;-0.00%;0.00%'


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    if widths:
        autofit(ws, widths)


def title_block(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")


def main():
    wb = Workbook()

    # ── Aba 1: Síntese ──
    ws = wb.active
    ws.title = "Síntese"
    title_block(ws, "Estudo 11: IBS total do Brasil, 2029-2033",
                "Série histórica (2013-2025) e projeção com PIB e inflação oficiais, memória de cálculo")

    p2029 = next(r for r in PROJ["projecao"] if r["ano"] == 2029)
    p2033 = next(r for r in PROJ["projecao"] if r["ano"] == 2033)

    rows = [
        ("Pergunta", "Quanto será o IBS total do Brasil entre 2029 e 2033, dado o cronograma constitucional de transição e projeções oficiais de PIB e inflação?"),
        ("", ""),
        ("Resposta", f"IBS bruto nacional: de R$ {p2029['ibs_bruto']/1e9:,.1f} bi em 2029 a R$ {p2033['ibs_bruto']/1e9:,.1f} bi em 2033 (R$ constantes de 2025)."),
        ("Arrecadação total em 2033 (ICMS+ISS+IBS)", f"R$ {p2033['bolo_projetado']/1e9:,.1f} bi ({p2033['bolo_pct_pib']:.2f}% do PIB, a mesma proporção de 2025, por premissa)"),
        ("", ""),
        ("Premissa central", "A receita de referência (ICMS+ISS+FECOP), como proporção do PIB (a carga tributária), é mantida constante no nível médio de 2024-2026 a partir de 2027: ela cresce sempre no mesmo ritmo da economia, nem mais nem menos rápido. É o próprio desenho legal da transição para o IBS (LC 214/2025, arts. 361 a 365, redação da LC 227/2026: o Senado fixa a alíquota de referência do IBS para manter essa razão constante)."),
        ("Fonte do PIB e inflação, 2026-2030", "Boletim Focus (Banco Central do Brasil), mediana das projeções de mercado, consultado via API do Sistema de Expectativas de Mercado."),
        ("Fonte do PIB e inflação, 2031-2033", "IFI, Instituição Fiscal Independente (Senado Federal), Relatório de Acompanhamento Fiscal (RAF) nº 107, 18/dez/2025: PIB real 2,2% a.a. (2027-2035), IPCA convergindo a 3,0% (centro da meta contínua)."),
        ("Fonte do ICMS", "SICONFI/STN, DCA Anexo I-C (2013-2025), comparado com o RREO Anexo 3 para 2015-2018 (diferença < 1,3%)."),
        ("Fonte do ISS", "SICONFI/STN, DCA Anexo I-C, todos os municípios brasileiros (2013-2025)."),
        ("Fonte do FECOP", "SICONFI/STN, DCA Anexo I-C (2019-2025); zero em 2013-2018; 2026 usa o valor de 2025 como estimativa preliminar. Mesmo critério do coeficiente validado no Estudo 06."),
        ("Base legal da transição", "ADCT arts. 128, 130, 131 e 132 (EC 132/2023); LC 227/2026, arts. 51 e 114 a 116 (CGIBS e critério histórico/destino)."),
        ("Cronograma f_a/s_a", "Fração remanescente de ICMS/ISS (f_a) e fração já convertida em IBS (s_a=1-f_a) a cada ano da transição, conforme ADCT arts. 128 e 131."),
        ("Divisão do IBS bruto", "Fração alpha_a distribuída pelo critério histórico; o restante pelo critério destino, líquido do CGIBS (art. 51 LC 227/2026) e do Seguro-Receita (5%, ADCT art. 132). Mesmos parâmetros já publicados no Estudo 06, usados aqui só para o agregado nacional (aba 'Projeção IBS 2029-2033')."),
        ("O que este arquivo mostra", "Aba 'Série histórica': dados de entrada SICONFI 2013-2025. Aba 'Premissas macro': dados de entrada Focus/IFI 2026-2033. Aba 'PIB projetado': fórmulas que compõem o PIB real ano a ano (R$ constantes de 2025, IPCA projetado não aplicado). Aba 'Projeção IBS': fórmulas que aplicam a carga tributária constante, o cronograma ADCT e a divisão do IBS bruto. Mude qualquer premissa nas abas de entrada (em azul) e a projeção recalcula."),
        ("Gerado em", date.today().isoformat()),
        ("Página do estudo", "https://www.eduardoreisaraujo.com.br/estudos/ibs-projecao-nacional.html"),
    ]
    r = 4
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = H_FONT if label else BODY_FONT
        vcell = ws.cell(row=r, column=2, value=value)
        vcell.font = BODY_FONT
        vcell.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1
    autofit(ws, [26, 20, 20, 20, 20, 20])
    for rr in range(4, r):
        ws.row_dimensions[rr].height = 30

    # ── Aba 2: Série histórica 2013-2025 (dados de entrada) ──
    ws2 = wb.create_sheet("Série histórica 2013-2025")
    title_block(ws2, "ICMS + ISS + FECOP, Brasil: série histórica 2013-2025",
                "Dados de entrada (células em azul), SICONFI/STN")
    headers = ["Ano", "ICMS (R$)", "ISS (R$)", "FECOP (R$)", "Total nominal (R$)", "Total real, 2025 (R$)",
               "PIB nominal (R$)", "Total / PIB", "Fonte ICMS"]
    write_header_row(ws2, 4, headers, widths=[8, 20, 20, 16, 22, 22, 20, 16, 12])
    hist_start_row = 5
    for i, h in enumerate(PROJ["historico"]):
        rr = hist_start_row + i
        ws2.cell(row=rr, column=1, value=h["ano"]).font = INPUT_FONT
        ws2.cell(row=rr, column=2, value=h["icms"]).font = INPUT_FONT
        ws2.cell(row=rr, column=3, value=h["iss"]).font = INPUT_FONT
        ws2.cell(row=rr, column=4, value=h.get("fecop", 0)).font = INPUT_FONT
        c5 = ws2.cell(row=rr, column=5, value=f"=B{rr}+C{rr}+D{rr}")
        c5.font = FORMULA_FONT
        ws2.cell(row=rr, column=6, value=h["bolo_real_2025"]).font = INPUT_FONT
        ws2.cell(row=rr, column=7, value=h["pib_nominal"]).font = INPUT_FONT
        c8 = ws2.cell(row=rr, column=8, value=f"=E{rr}/G{rr}")
        c8.font = FORMULA_FONT
        c8.number_format = PCT_FMT
        ws2.cell(row=rr, column=9, value=h["fonte_icms"]).font = BODY_FONT
        for col in (2, 3, 4, 5, 6, 7):
            ws2.cell(row=rr, column=col).number_format = RS_FMT_FULL
        for col in range(1, 10):
            ws2.cell(row=rr, column=col).border = BORDER
    hist_end_row = hist_start_row + len(PROJ["historico"]) - 1
    note_row = hist_end_row + 2
    ws2.cell(row=note_row, column=1,
             value="ICMS e ISS 2013-2025, FECOP 2019-2025: SICONFI/STN, DCA Anexo I-C, todos os estados e municípios. O SICONFI não publica o DCA "
                   "antes de 2013 (FECOP tratado como zero em 2013-2018, sem coleta equivalente via RREO). Para 2015-2018, o ICMS do DCA foi comparado "
                   "com o RREO Anexo 3 (conta ICMSLiquidoExcetoTransferenciasEFUNDEB, coluna \"TOTAL últimos 12 meses\"): diferença < 1,3% em todos os anos. "
                   "FECOP incluído porque a legislação da reforma o trata como parte da base substituída pelo IBS nos estados que o cobram (mesmo critério "
                   "do coeficiente validado no Estudo 06).").font = NOTE_FONT
    ws2.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)
    ws2.cell(row=note_row, column=1).alignment = WRAP
    ws2.row_dimensions[note_row].height = 40

    # ── Aba 3: Premissas macro 2026-2033 (dados de entrada) ──
    ws3 = wb.create_sheet("Premissas macro 2026-2033")
    title_block(ws3, "Crescimento do PIB e inflação, 2026-2033",
                "Dados de entrada (células em azul), Boletim Focus (BCB) e IFI (RAF 107)")
    headers3 = ["Ano", "PIB real (%)", "IPCA (%)", "Fonte"]
    write_header_row(ws3, 4, headers3, widths=[8, 14, 14, 42])
    macro_start_row = 5
    for i, m in enumerate(PROJ["macro_path"]):
        rr = macro_start_row + i
        ws3.cell(row=rr, column=1, value=m["ano"]).font = INPUT_FONT
        c2 = ws3.cell(row=rr, column=2, value=m["pib_real_pct"])
        c2.font = INPUT_FONT
        c2.number_format = PCT_FMT_SIGNED
        c3 = ws3.cell(row=rr, column=3, value=m["ipca_pct"])
        c3.font = INPUT_FONT
        c3.number_format = PCT_FMT_SIGNED
        ws3.cell(row=rr, column=4, value=m["fonte"]).font = BODY_FONT
        for col in range(1, 5):
            ws3.cell(row=rr, column=col).border = BORDER
    macro_end_row = macro_start_row + len(PROJ["macro_path"]) - 1
    note_row3 = macro_end_row + 2
    ws3.cell(row=note_row3, column=1,
             value="2026-2030: mediana do Boletim Focus (BCB), consultada via API do Sistema de Expectativas de Mercado. "
                   "2031-2033: fora do horizonte do Focus (~5 anos); usa-se a projeção da IFI (RAF 107, 18/dez/2025), constante nos três anos por simplificação, "
                   "já que a IFI projeta uma taxa média para 2027-2035, não ano a ano.").font = NOTE_FONT
    ws3.merge_cells(start_row=note_row3, start_column=1, end_row=note_row3, end_column=4)
    ws3.cell(row=note_row3, column=1).alignment = WRAP
    ws3.row_dimensions[note_row3].height = 55

    # ── Aba 4: PIB projetado (fórmulas) ──
    ws4 = wb.create_sheet("PIB projetado")
    title_block(ws4, "PIB real projetado, 2025-2033 (R$ constantes de 2025)",
                "Fórmulas: PIB(t) = PIB(t-1) × (1 + PIB real). Não compõe IPCA -- a partir de "
                "ago/2026 este estudo projeta em R$ constantes de 2025, não em R$ nominais.")
    headers4 = ["Ano", "PIB real, R$ 2025 (R$)", "Crescimento real (%)"]
    write_header_row(ws4, 4, headers4, widths=[8, 22, 20])
    ano2025 = next(h for h in PROJ["historico"] if h["ano"] == 2025)
    rr = 5
    ws4.cell(row=rr, column=1, value=2025).font = INPUT_FONT
    ws4.cell(row=rr, column=2, value=ano2025["pib_nominal"]).font = INPUT_FONT
    ws4.cell(row=rr, column=2).number_format = RS_FMT_FULL
    ws4.cell(row=rr, column=3, value=None)
    for col in range(1, 4):
        ws4.cell(row=rr, column=col).border = BORDER
    pib2025_row = rr
    prev_row = rr
    for i, m in enumerate(PROJ["macro_path"]):
        rr = 6 + i
        macro_row = macro_start_row + i
        ws4.cell(row=rr, column=1, value=f"=Ano('Premissas macro 2026-2033'!A{macro_row})" if False else m["ano"]).font = FORMULA_FONT
        cpib = ws4.cell(row=rr, column=2,
                         value=f"=B{prev_row}*(1+'Premissas macro 2026-2033'!B{macro_row})")
        cpib.font = FORMULA_FONT
        cpib.number_format = RS_FMT_FULL
        cgrow = ws4.cell(row=rr, column=3, value=f"=B{rr}/B{prev_row}-1")
        cgrow.font = FORMULA_FONT
        cgrow.number_format = PCT_FMT
        for col in range(1, 4):
            ws4.cell(row=rr, column=col).border = BORDER
        prev_row = rr
    pib_rows_by_year = {2025: pib2025_row}
    for i, m in enumerate(PROJ["macro_path"]):
        pib_rows_by_year[m["ano"]] = 6 + i

    # ── Aba 5: Projeção IBS 2029-2033 (fórmulas) ──
    ws5 = wb.create_sheet("Projeção IBS 2029-2033")
    title_block(ws5, "Projeção do IBS total, Brasil, 2029-2033",
                "Fórmulas: Receita(t) = Razão(2024-2026) × PIB(t); ICMS+ISS residual = Receita × fa; IBS bruto = Receita × sa")

    hist_row_by_year = {h["ano"]: hist_start_row + i for i, h in enumerate(PROJ["historico"])}
    row2024, row2025 = hist_row_by_year[2024], hist_row_by_year[2025]
    ref2026 = PROJ["_meta"]["receita_referencia"]["anos"][2]  # ano 2026, estimativa preliminar

    ws5["A4"] = "Receita de referência (LC 214/2025, arts. 361-365): média da razão receita/PIB em 2024-2026"
    ws5["A4"].font = H_FONT
    ws5.merge_cells("A4:H4")

    headers_ref = ["Ano", "ICMS+ISS+FECOP (R$)", "PIB (R$)", "Razão", "Situação"]
    write_header_row(ws5, 5, headers_ref, widths=[8, 20, 14, 20, 20, 22, 20, 14])
    ws5.cell(row=6, column=1, value=2024).font = FORMULA_FONT
    ws5.cell(row=6, column=2, value=f"='Série histórica 2013-2025'!E{row2024}").font = FORMULA_FONT
    ws5.cell(row=6, column=3, value=f"='Série histórica 2013-2025'!G{row2024}").font = FORMULA_FONT
    ws5.cell(row=6, column=4, value="=B6/C6").font = FORMULA_FONT
    ws5.cell(row=6, column=5, value="fechado (DCA)").font = BODY_FONT
    ws5.cell(row=7, column=1, value=2025).font = FORMULA_FONT
    ws5.cell(row=7, column=2, value=f"='Série histórica 2013-2025'!E{row2025}").font = FORMULA_FONT
    ws5.cell(row=7, column=3, value=f"='Série histórica 2013-2025'!G{row2025}").font = FORMULA_FONT
    ws5.cell(row=7, column=4, value="=B7/C7").font = FORMULA_FONT
    ws5.cell(row=7, column=5, value="fechado (DCA)").font = BODY_FONT
    ws5.cell(row=8, column=1, value=2026).font = INPUT_FONT
    ws5.cell(row=8, column=2, value=ref2026["bolo"]).font = INPUT_FONT
    ws5.cell(row=8, column=3, value=ref2026["pib"]).font = INPUT_FONT
    ws5.cell(row=8, column=4, value="=B8/C8").font = FORMULA_FONT
    ws5.cell(row=8, column=5, value="estimativa preliminar (RREO 12m + PIB Focus)").font = BODY_FONT
    for rr in (6, 7, 8):
        ws5.cell(row=rr, column=2).number_format = RS_FMT_FULL
        ws5.cell(row=rr, column=3).number_format = RS_FMT_FULL
        ws5.cell(row=rr, column=4).number_format = PCT_FMT
        for col in range(1, 6):
            ws5.cell(row=rr, column=col).border = BORDER
    ws5.cell(row=9, column=1, value="Média 2024-2026 (usada na projeção)").font = TOTAL_FONT
    ws5.merge_cells(start_row=9, start_column=1, end_row=9, end_column=3)
    ws5.cell(row=9, column=4, value="=AVERAGE(D6:D8)").font = TOTAL_FONT
    ws5.cell(row=9, column=4).number_format = PCT_FMT
    ws5.cell(row=9, column=4).fill = TOTAL_FILL
    ws5.cell(row=9, column=1).fill = TOTAL_FILL
    for col in range(1, 6):
        ws5.cell(row=9, column=col).border = BORDER
    razao_cell = "$D$9"

    ws5.cell(row=11, column=1,
             value="O Teto de Referência (ADCT art. 130, §4º e §5º) compara a média 2029-2033 de CBS+Imposto "
                   "Seletivo+IBS com a média 2012-2021 de IPI+ICMS+ISS+PIS/Cofins+IOF-seguros. Não altera os "
                   "valores desta planilha: mesmo se ultrapassado, a redução só vale a partir de 2035 (fora "
                   "desta projeção), e exige dados federais fora do escopo deste estudo (só ICMS/ISS/IBS).").font = NOTE_FONT
    ws5.merge_cells(start_row=11, start_column=1, end_row=11, end_column=8)
    ws5.cell(row=11, column=1).alignment = WRAP
    ws5.row_dimensions[11].height = 45

    headers5 = ["Ano", "f_a (ICMS+ISS residual)", "s_a (IBS)", "PIB real, R$ 2025 (R$)",
                "Receita projetada (R$)", "ICMS+ISS residual (R$)", "IBS bruto (R$)", "Carga tributária (% PIB)"]
    write_header_row(ws5, 13, headers5, widths=[8, 20, 14, 20, 20, 22, 20, 14])
    proj_start_row = 14
    for i, p in enumerate(PROJ["projecao"]):
        rr = proj_start_row + i
        pib_row = pib_rows_by_year[p["ano"]]
        ws5.cell(row=rr, column=1, value=p["ano"]).font = FORMULA_FONT
        cfa = ws5.cell(row=rr, column=2, value=p["fa"])
        cfa.font = INPUT_FONT
        cfa.number_format = PCT_FMT
        csa = ws5.cell(row=rr, column=3, value=f"=1-B{rr}")
        csa.font = FORMULA_FONT
        csa.number_format = PCT_FMT
        cpib = ws5.cell(row=rr, column=4, value=f"='PIB projetado'!B{pib_row}")
        cpib.font = FORMULA_FONT
        cpib.number_format = RS_FMT_FULL
        cbolo = ws5.cell(row=rr, column=5, value=f"={razao_cell}*D{rr}")
        cbolo.font = FORMULA_FONT
        cbolo.number_format = RS_FMT_FULL
        cres = ws5.cell(row=rr, column=6, value=f"=E{rr}*B{rr}")
        cres.font = FORMULA_FONT
        cres.number_format = RS_FMT_FULL
        cibs = ws5.cell(row=rr, column=7, value=f"=E{rr}*C{rr}")
        cibs.font = FORMULA_FONT
        cibs.number_format = RS_FMT_FULL
        cratio = ws5.cell(row=rr, column=8, value=f"=E{rr}/D{rr}")
        cratio.font = FORMULA_FONT
        cratio.number_format = PCT_FMT
        for col in range(1, 9):
            ws5.cell(row=rr, column=col).border = BORDER

    note_row5 = proj_start_row + len(PROJ["projecao"]) + 1
    ws5.cell(row=note_row5, column=1,
             value="f_a (fração remanescente de ICMS+ISS) e s_a=1-f_a (fração já convertida em IBS) seguem o cronograma da transição "
                   "constitucional (ADCT arts. 128 e 131, LC 227/2026). \"IBS bruto\" é a arrecadação total antes de qualquer dedução; "
                   "a divisão entre critério histórico, critério destino, CGIBS e Seguro-Receita está na seção abaixo.").font = NOTE_FONT
    ws5.merge_cells(start_row=note_row5, start_column=1, end_row=note_row5, end_column=8)
    ws5.cell(row=note_row5, column=1).alignment = WRAP
    ws5.row_dimensions[note_row5].height = 45

    # ── Seção: divisão do IBS bruto (critério histórico, destino, CGIBS, Seguro-Receita) ──
    div_title_row = note_row5 + 2
    ws5.cell(row=div_title_row, column=1,
             value="Divisão do IBS bruto (ADCT arts. 131-132; LC 227/2026 arts. 51 e 114-116): mesmos parâmetros alpha_a e c_a do Estudo 06").font = H_FONT
    ws5.merge_cells(start_row=div_title_row, start_column=1, end_row=div_title_row, end_column=9)

    headers_div = ["Ano", "alpha_a (histórico)", "c_a (CGIBS)", "rho (Seguro-Receita)",
                   "IBS histórico (R$)", "CGIBS (R$)", "Seguro-Receita (R$)", "IBS destino, líquido (R$)", "Total (confere IBS bruto)"]
    div_header_row = div_title_row + 1
    write_header_row(ws5, div_header_row, headers_div, widths=[8, 20, 14, 20, 20, 22, 20, 14, 22])
    div_start_row = div_header_row + 1
    for i, p in enumerate(PROJ["projecao"]):
        rr = div_start_row + i
        proj_row = proj_start_row + i
        ws5.cell(row=rr, column=1, value=p["ano"]).font = FORMULA_FONT
        calpha = ws5.cell(row=rr, column=2, value=p["alpha_a"])
        calpha.font = INPUT_FONT
        calpha.number_format = PCT_FMT
        cca = ws5.cell(row=rr, column=3, value=p["ca"])
        cca.font = INPUT_FONT
        cca.number_format = PCT_FMT
        crho = ws5.cell(row=rr, column=4, value=0.05)
        crho.font = INPUT_FONT
        crho.number_format = PCT_FMT
        chist = ws5.cell(row=rr, column=5, value=f"=G{proj_row}*B{rr}")
        chist.font = FORMULA_FONT
        chist.number_format = RS_FMT_FULL
        ccgibs = ws5.cell(row=rr, column=6, value=f"=G{proj_row}*(1-B{rr})*C{rr}")
        ccgibs.font = FORMULA_FONT
        ccgibs.number_format = RS_FMT_FULL
        cseguro = ws5.cell(row=rr, column=7, value=f"=G{proj_row}*(1-B{rr})*(1-C{rr})*D{rr}")
        cseguro.font = FORMULA_FONT
        cseguro.number_format = RS_FMT_FULL
        cdest = ws5.cell(row=rr, column=8, value=f"=G{proj_row}*(1-B{rr})*(1-C{rr})*(1-D{rr})")
        cdest.font = FORMULA_FONT
        cdest.number_format = RS_FMT_FULL
        ctotal = ws5.cell(row=rr, column=9, value=f"=E{rr}+F{rr}+G{rr}+H{rr}")
        ctotal.font = FORMULA_FONT
        ctotal.number_format = RS_FMT_FULL
        for col in range(1, 10):
            ws5.cell(row=rr, column=col).border = BORDER

    note_row_div = div_start_row + len(PROJ["projecao"]) + 1
    ws5.cell(row=note_row_div, column=1,
             value="alpha_a = fração do IBS bruto distribuída pelo critério histórico; (1-alpha_a) vai pelo critério destino, mas antes sofre "
                   "duas deduções, nesta ordem: c_a (CGIBS, art. 51 LC 227/2026) e rho=5% (Seguro-Receita, ADCT art. 132). As quatro colunas de "
                   "reais somam exatamente o IBS bruto do ano (coluna \"Total\" confere isso). alpha_a e c_a são os mesmos parâmetros já "
                   "publicados no Estudo 06 (estudos/ibs-projecao-arrecadacao-br.html), usados aqui só para mostrar o tamanho de cada fatia no "
                   "agregado nacional, sem entrar em qual estado ou município recebe o quê.").font = NOTE_FONT
    ws5.merge_cells(start_row=note_row_div, start_column=1, end_row=note_row_div, end_column=9)
    ws5.cell(row=note_row_div, column=1).alignment = WRAP
    ws5.row_dimensions[note_row_div].height = 65

    wb.save(OUTPUT)
    print(f"Gravado em {OUTPUT}")


if __name__ == "__main__":
    main()
