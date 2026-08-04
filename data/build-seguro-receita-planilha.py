#!/usr/bin/env python3
"""
Planilha auditável do Seguro-Receita: reproduz, com fórmulas do Excel
visíveis (não valores colados), o algoritmo de nivelamento sequencial do
art. 117 da LC 227/2026 sobre os 5.596 entes (27 estados + DF + 5.569
municípios), para 2029-2033.

Método (equivalente ao "water_fill" por busca binária de
build-seguro-receita-repasses.py, verificado numericamente contra ele --
mesmo nível L, mesma soma, mesmos beneficiários, diferença < R$ 0,01):
o "nivelamento por enchimento de água" pode ser resolvido, em vez de por
busca numérica, do jeito que um humano faria com papel e caneta --
ordenar do pior pro melhor e ir somando:

  1. Ordena os entes por razão (numerador/denominador) crescente.
  2. Acumula o denominador (D) e o numerador (N) dos entes já somados.
  3. Para cada ente k, calcula quanto custaria elevar TODOS os entes
     antes dele até o nível da razão de k: custo_k = razão_k × D_(k-1) − N_(k-1).
     Essa coluna só cresce (é uma soma acumulada de incrementos não
     negativos), então dá pra achar exatamente onde o fundo se esgota.
  4. Conta quantos entes têm custo_k ≤ fundo -- esse é o último ente
     TOTALMENTE alcançado antes de o fundo acabar no meio do caminho.
  5. O nível final L é o ponto exato, entre a razão desse último ente e a
     do próximo, onde o fundo se esgota.
  6. Repasse de cada ente = MAX(0, L × denominador − numerador).

Cada uma dessas contas é uma fórmula independente na planilha, célula por
célula -- dá pra abrir, seguir a lógica linha a linha, e conferir contra
os números já publicados no Estudo 12.

Uso:
  python3 build-seguro-receita-planilha.py
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
SRC = HERE / "seguro-receita-repasses.json"
OUT = HERE.parent / "materiais" / "seguro-receita-repasses-memoria-calculo.xlsx"

ANOS = [2029, 2030, 2031, 2032, 2033]

HEADER_FILL = PatternFill(start_color="1D3557", end_color="1D3557", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BOLD = Font(bold=True)
MONEY_FMT = '#,##0.00'
PCT_FMT = '0.0000%'


def main():
    with open(SRC) as f:
        data = json.load(f)

    # Ordena pela razão de 2033 -- a ordem é (matematicamente) a mesma em
    # todos os anos, porque numerador e fundo crescem na mesma proporção
    # ano a ano; só o nível L muda de tamanho, não a fila.
    ents = sorted(data["anos"]["2033"]["entidades"], key=lambda e: e["razao"])
    ents_por_ano = {ano: {e["id"]: e for e in data["anos"][str(ano)]["entidades"]} for ano in ANOS}
    n = len(ents)

    wb = Workbook()

    # ── Aba "Leia-me" ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Leia-me"
    ws.column_dimensions["A"].width = 100
    linhas = [
        "Planilha auditável: repasses do Seguro-Receita por ente, 2029-2033",
        "",
        "Base legal: ADCT art. 132 (EC 132/2023) e LC 227/2026 art. 117.",
        "Site: eduardoreisaraujo.com.br/estudos/seguro-receita-repasses.html (Estudo 12)",
        "",
        "COMO LER A ABA 'Cálculo'",
        "",
        "Cada linha é um ente (estado, DF ou município), ordenado da MENOR razão",
        "para a MAIOR -- ou seja, de quem está pior para quem está melhor nesse",
        "critério específico. As colunas, para cada ano, são todas fórmulas vivas:",
        "",
        "  numerador_AAAA   = quanto o ente recebe hoje pelo critério destino no ano",
        "  denom_capado     = receita média de referência do ente, já com o teto de",
        "                     3x a média nacional per capita aplicado (não muda por ano)",
        "  razão_AAAA       = numerador_AAAA / denom_capado",
        "  D_acumulado      = soma do denom_capado de todos os entes até esta linha",
        "                     (inclusive) -- não muda por ano",
        "  N_acumulado_AAAA = soma do numerador_AAAA de todos os entes até esta linha",
        "  custo_AAAA       = razão_AAAA × D_acumulado(linha anterior) − N_acumulado(linha anterior)",
        "                     -- quanto custaria elevar TODOS os entes antes deste até",
        "                     a razão deste ente. Essa coluna só cresce, linha a linha.",
        "  repasse_AAAA     = MAX(0, L_AAAA × denom_capado − numerador_AAAA)",
        "",
        "O nível L de cada ano (aba 'Resumo') é achado assim: conta quantas linhas",
        "têm custo_AAAA <= fundo do ano (COUNTIF) -- esse número de linhas (m) marca",
        "o último ente totalmente alcançado antes do fundo se esgotar. O nível final é",
        "",
        "  L = razão(linha m) + (fundo − custo(linha m)) / D_acumulado(linha m)",
        "",
        "Confira: a soma de todos os repasse_AAAA bate exatamente com o fundo do ano",
        "(célula de conferência na aba Resumo). Se você mudar um valor de entrada",
        "(numerador ou denom_capado) linha por linha, todas as fórmulas recalculam --",
        "inclusive quem entra ou sai da lista de beneficiados.",
        "",
        "EXEMPLO PEQUENO (mesma lógica, só pra ilustrar antes de olhar as 5.596 linhas reais)",
        "",
        "4 entes fictícios, fundo de R$ 10 milhões:",
        "  A: numerador R$ 4mi, denominador R$ 100mi -> razão 4,0%",
        "  B: numerador R$ 4,8mi, denominador R$ 80mi -> razão 6,0%",
        "  C: numerador R$ 3mi, denominador R$ 50mi -> razão 6,0%",
        "  D: numerador R$ 30mi, denominador R$ 200mi -> razão 15,0%",
        "",
        "Ordem: A (pior) < B = C (empatados) < D (melhor, fora do alcance do fundo).",
        "Passo 1: eleva A de 4% para 6% (nível de B/C) -- custa R$ 2mi. Sobra R$ 8mi.",
        "Passo 2: A, B e C sobem juntos; juntos precisariam de R$ 20,7mi pra alcançar",
        "D (15%), mas só sobram R$ 8mi -- o fundo acaba num nível comum de 9,48%.",
        "Resultado: A recebe R$ 5,48mi; B recebe R$ 2,78mi; C recebe R$ 1,74mi",
        "(os três empatam em 9,48%); D não recebe nada. Soma = R$ 10mi, o fundo inteiro.",
        "Ver a aba 'Exemplo' para essa mesma conta com fórmulas.",
        "",
        "LIMITAÇÕES (ver metodologia completa na página do Estudo 12)",
        "",
        "- Estimativa própria, não o cálculo oficial do CGIBS.",
        "- Anual em vez de mensal (a lei apura uma média móvel de 12 meses).",
        "- População: média de 2019, 2020, 2021, 2024 e 2025 (IBGE) -- 2022-2023 não",
        "  têm estimativa normal (Censo Demográfico a substituiu) e 2026 não foi",
        "  publicado ainda; a lei pede a média entre 2019 e 2026.",
        "- DF: numerador restrito ao ICMS, por consistência com o Estudo 06.",
    ]
    for i, texto in enumerate(linhas, start=1):
        c = ws.cell(row=i, column=1, value=texto)
        if i == 1:
            c.font = Font(bold=True, size=13)
        elif texto.isupper() or (texto.endswith(("'Cálculo'", "'Resumo')", "olhar as 5.596 linhas reais)"))):
            c.font = BOLD

    # ── Aba "Exemplo" (o exemplo pequeno, com fórmulas de verdade) ──────
    we = wb.create_sheet("Exemplo")
    exemplo = [
        ("A", 4_000_000, 100_000_000),
        ("B", 4_800_000, 80_000_000),
        ("C", 3_000_000, 50_000_000),
        ("D", 30_000_000, 200_000_000),
    ]
    headers_ex = ["Ente", "Numerador (R$)", "Denominador (R$)", "Razão", "D acumulado", "N acumulado", "Custo p/ chegar aqui", "Repasse (R$)"]
    for j, h in enumerate(headers_ex, start=1):
        cell = we.cell(row=1, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for i, (nome, numerador, denom) in enumerate(exemplo, start=2):
        we.cell(row=i, column=1, value=nome)
        we.cell(row=i, column=2, value=numerador).number_format = MONEY_FMT
        we.cell(row=i, column=3, value=denom).number_format = MONEY_FMT
        we.cell(row=i, column=4, value=f"=B{i}/C{i}").number_format = PCT_FMT
        we.cell(row=i, column=5, value=f"=C{i}" if i == 2 else f"=E{i-1}+C{i}").number_format = MONEY_FMT
        we.cell(row=i, column=6, value=f"=B{i}" if i == 2 else f"=F{i-1}+B{i}").number_format = MONEY_FMT
        we.cell(row=i, column=7, value=f"=D{i}*E{i-1}-F{i-1}" if i > 2 else "=0").number_format = MONEY_FMT
    we.cell(row=6, column=1, value="Fundo (R$)").font = BOLD
    we.cell(row=6, column=2, value=10_000_000).number_format = MONEY_FMT
    we.cell(row=7, column=1, value="m (linhas totalmente cobertas)").font = BOLD
    we.cell(row=7, column=2, value="=COUNTIF(G2:G5,\"<=\"&B6)")
    we.cell(row=8, column=1, value="L (nível final)").font = BOLD
    we.cell(row=8, column=2, value="=INDEX(D2:D5,B7)+(B6-INDEX(G2:G5,B7))/INDEX(E2:E5,B7)").number_format = PCT_FMT
    for i in range(2, 6):
        we.cell(row=i, column=8, value=f"=MAX(0,$B$8*C{i}-B{i})").number_format = MONEY_FMT
    we.cell(row=9, column=1, value="Soma dos repasses (deve bater com o fundo)").font = BOLD
    we.cell(row=9, column=2, value="=SUM(H2:H5)").number_format = MONEY_FMT
    for col, width in zip("ABCDEFGH", [8, 16, 18, 10, 16, 16, 18, 16]):
        we.column_dimensions[col].width = width

    # ── Aba "Cálculo" (5.596 linhas, fórmulas vivas) ────────────────────
    wc = wb.create_sheet("Cálculo")
    headers = ["Ente", "Esfera", "UF", "denom_capado", "D_acumulado"]
    for ano in ANOS:
        headers += [f"numerador_{ano}", f"razão_{ano}", f"N_acumulado_{ano}", f"custo_{ano}", f"repasse_{ano}"]
    for j, h in enumerate(headers, start=1):
        cell = wc.cell(row=1, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    wc.freeze_panes = "A2"

    COL_DENOM = 4
    COL_DACUM = 5
    for i, e in enumerate(ents, start=2):
        wc.cell(row=i, column=1, value=e["nome"])
        wc.cell(row=i, column=2, value="Estado" if e["esfera"] == "estado" else "Município")
        wc.cell(row=i, column=3, value=e["uf"])
        wc.cell(row=i, column=COL_DENOM, value=e["denom_capado"]).number_format = MONEY_FMT
        dcell = wc.cell(row=i, column=COL_DACUM)
        dcell.value = f"=D{i}" if i == 2 else f"=E{i-1}+D{i}"
        dcell.number_format = MONEY_FMT

        col = 6
        for ano in ANOS:
            ea = ents_por_ano[ano][e["id"]]
            col_num, col_raz, col_nacum, col_custo, col_repasse = col, col + 1, col + 2, col + 3, col + 4
            L_num, L_raz, L_nacum, L_custo = (get_column_letter(c) for c in (col_num, col_raz, col_nacum, col_custo))

            wc.cell(row=i, column=col_num, value=ea["numerador"]).number_format = MONEY_FMT
            c_raz = wc.cell(row=i, column=col_raz, value=f"={L_num}{i}/$D{i}")
            c_raz.number_format = PCT_FMT
            c_nacum = wc.cell(row=i, column=col_nacum,
                               value=f"={L_num}{i}" if i == 2 else f"={L_nacum}{i-1}+{L_num}{i}")
            c_nacum.number_format = MONEY_FMT
            custo_formula = "=0" if i == 2 else f"={L_raz}{i}*$E{i-1}-{L_nacum}{i-1}"
            wc.cell(row=i, column=col_custo, value=custo_formula).number_format = MONEY_FMT
            wc.cell(row=i, column=col_repasse,
                    value=f"=MAX(0,Resumo!$D${3 + ANOS.index(ano)}*$D{i}-{L_num}{i})").number_format = MONEY_FMT
            col += 5

    for col, width in zip("ABC", [26, 12, 6]):
        wc.column_dimensions[col].width = width
    for j in range(4, 4 + 1 + 5 * 5):
        wc.column_dimensions[get_column_letter(j)].width = 14

    # ── Aba "Resumo" (fundo, m, L, checagem por ano; top beneficiários) ─
    # Layout fixo: linha 1 = título, linha 2 = cabeçalhos, linhas 3-7 = 2029-2033.
    # As fórmulas de repasse da aba "Cálculo" referenciam Resumo!$D$3..$D$7 (o nível L
    # de cada ano), então essas linhas não podem se mover.
    wr = wb.create_sheet("Resumo")
    wr.cell(row=1, column=1, value="RESUMO POR ANO").font = Font(bold=True, size=13)
    headers_resumo = ["Ano", "Fundo (R$)", "m (linhas cobertas)", "L (nível final)",
                       "Soma dos repasses (Cálculo)", "Confere com o fundo?"]
    for j, h in enumerate(headers_resumo, start=1):
        wr.cell(row=2, column=j, value=h).font = BOLD
    for idx, ano in enumerate(ANOS):
        row = 3 + idx
        pool = data["anos"][str(ano)]["pool"]
        col_custo_letra = get_column_letter(6 + idx * 5 + 3)  # coluna custo_AAAA na aba Cálculo
        col_raz_letra = get_column_letter(6 + idx * 5 + 1)
        col_dacum_letra = "E"
        wr.cell(row=row, column=1, value=ano)
        wr.cell(row=row, column=2, value=pool).number_format = MONEY_FMT
        wr.cell(row=row, column=3,
                value=f"=COUNTIF(Cálculo!{col_custo_letra}2:{col_custo_letra}{1+n},\"<=\"&B{row})")
        wr.cell(row=row, column=4,
                value=(f"=INDEX(Cálculo!{col_raz_letra}2:{col_raz_letra}{1+n},C{row})"
                       f"+(B{row}-INDEX(Cálculo!{col_custo_letra}2:{col_custo_letra}{1+n},C{row}))"
                       f"/INDEX(Cálculo!{col_dacum_letra}2:{col_dacum_letra}{1+n},C{row})")).number_format = '0.000000%'
        col_repasse_letra = get_column_letter(6 + idx * 5 + 4)
        wr.cell(row=row, column=5,
                value=f"=SUM(Cálculo!{col_repasse_letra}2:{col_repasse_letra}{1+n})").number_format = MONEY_FMT
        wr.cell(row=row, column=6, value=f'=IF(ABS(E{row}-B{row})<1,"Sim","NÃO -- checar")')
    for col, width in zip("ABCDEF", [8, 18, 16, 16, 20, 16]):
        wr.column_dimensions[col].width = width

    # Move "Resumo" para antes de "Cálculo" (fórmulas da aba Cálculo referenciam Resumo!$D$)
    wb.move_sheet("Resumo", offset=-1)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Salvo em {OUT}")
    print(f"Linhas na aba Cálculo: {n}")


if __name__ == "__main__":
    main()
