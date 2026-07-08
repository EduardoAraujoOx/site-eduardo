#!/usr/bin/env python3
"""
Gera a planilha compartilhável do Estudo 08 (cota-parte do ICMS aos municípios do ES)
a partir de data/cota-parte-municipal-es.json.

Saída: materiais/cota-parte-icms-municipios-es.xlsx, com abas:
  - Síntese: pergunta, achados principais e reconciliação agregada por ano
  - Metodologia: resumo das fontes e da correção bruto/líquido
  - 2023, 2024, 2025: oficial (SEFAZ) x declarado (SICONFI), um município por linha
  - Validação interna: DCA x SIOPE x SIOPS, 2019-2025

Uso:
  python3 data/build-cota-parte-xlsx.py
"""

import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
DATA = json.loads((HERE / "cota-parte-municipal-es.json").read_text())
OUTPUT = HERE.parent / "materiais" / "cota-parte-icms-municipios-es.xlsx"

NAVY = "1A3A5C"
LIGHT_BG = "EDF2F7"
WHITE = "FFFFFF"

TITLE_FONT = Font(name="Calibri", size=15, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="6E6E6E")
H_FONT = Font(name="Calibri", size=11, bold=True, color="1D1D1B")
BODY_FONT = Font(name="Calibri", size=10.5, color="1D1D1B")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Calibri", size=10.5, bold=True, color=WHITE)
TOTAL_FILL = PatternFill("solid", fgColor=LIGHT_BG)
TOTAL_FONT = Font(name="Calibri", size=10.5, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
RS_FMT = '"R$" #,##0.00'
PCT_FMT = '+0.000%;-0.000%;0.000%'


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, headers, widths=None):
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def build_sintese(wb):
    ws = wb.active
    ws.title = "Síntese"
    autofit(ws, [26, 20, 20, 18, 18])

    ws.merge_cells("A1:E1")
    ws["A1"] = "Cota-parte do ICMS aos municípios do Espírito Santo: declarado (SICONFI) x oficial (SEFAZ-ES)"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 34
    ws["A1"].alignment = WRAP

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Síntese do levantamento · Eduardo Reis Araújo · gerado em {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font = SUBTITLE_FONT

    r = 4
    ws.cell(row=r, column=1, value="Pergunta").font = H_FONT
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value=(
        "A cota-parte do ICMS que os 78 municípios do Espírito Santo declaram ao SICONFI (DCA), "
        "ao SIOPE e ao SIOPS bate com o valor que a SEFAZ-ES efetivamente lhes repassou? Divergências "
        "nessas declarações podem distorcer o coeficiente de transição de participação federativa da "
        "reforma tributária (art. 115 da LC 227/2026), que depende de receitas autodeclaradas pelos entes."
    )).alignment = WRAP
    ws.row_dimensions[r].height = 60
    r += 2

    ws.cell(row=r, column=1, value="Principal achado").font = H_FONT
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value=(
        "A coluna 'Cota-Parte do ICMS' da base declarada corresponde ao valor BRUTO (antes da retenção "
        "de 20% ao FUNDEB). Comparado ao ICMS bruto oficial da SEFAZ-ES, o agregado estadual reconcilia "
        "quase integralmente em 2023-2025 (desvio de 0,01 a 0,07 pontos percentuais), o que confirma a "
        "correspondência entre as duas fontes. As divergências relevantes ocorrem no nível municipal, "
        "com casos individuais de até ~6%, detalhados nas abas 2023, 2024 e 2025 desta planilha."
    )).alignment = WRAP
    ws.row_dimensions[r].height = 75
    r += 2

    ws.cell(row=r, column=1, value="Reconciliação agregada, Espírito Santo (ICMS bruto)").font = H_FONT
    r += 1
    header_row = r
    write_header_row(ws, header_row, ["Ano", "Oficial (SEFAZ)", "Declarado (SICONFI)", "Diferença (R$)", "Diferença (%)"])
    r += 1
    for ano in ("2023", "2024", "2025"):
        a = DATA["agregado_es_por_ano"][ano]
        ws.cell(row=r, column=1, value=int(ano))
        ws.cell(row=r, column=2, value=a["oficial_icms_bruto"]).number_format = RS_FMT
        ws.cell(row=r, column=3, value=a["declarado_siconfi"]).number_format = RS_FMT
        ws.cell(row=r, column=4, value=a["diff_rs"]).number_format = RS_FMT
        ws.cell(row=r, column=5, value=a["diff_pct"] / 100).number_format = PCT_FMT
        r += 1
    ws.freeze_panes = None  # only meaningful per-sheet on the detail tabs
    r += 1

    ws.cell(row=r, column=1, value="Outras checagens de qualidade (2019-2025)").font = H_FONT
    r += 1
    write_header_row(ws, r, ["Checagem", "Resultado", "", "", ""])
    r += 1
    checks = [
        ("Dados faltantes (DCA/SIOPE/SIOPS em branco)", f"{len(DATA['qualidade_dados_faltantes'])} registros, todos de 2025 (defasagem de entrega ao SIOPE)"),
        ("Dados duplicados (mesmo valor entre municípios distintos)", f"{len(DATA['qualidade_dados_duplicados'])} casos encontrados entre os 78 municípios"),
        ("Saltos atípicos (subida > 50% revertida > 30% no ano seguinte)", f"{len(DATA['qualidade_saltos_atipicos'])} casos, todos em ISS de municípios pequenos"),
    ]
    for label, value in checks:
        ws.cell(row=r, column=1, value=label)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.cell(row=r, column=2, value=value)
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Fontes").font = H_FONT
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value=DATA["fonte_oficial"]).alignment = WRAP
    ws.row_dimensions[r].height = 30
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value=DATA["fonte_declarado"]).alignment = WRAP
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value=(
        "Estudo completo, metodologia, gráficos e demais tabelas: "
        "https://www.eduardoreisaraujo.com.br/estudos/cota-parte-icms-municipios-es.html"
    )).font = Font(name="Calibri", size=10.5, color=NAVY, underline="single")

    for row in ws.iter_rows(min_row=1, max_row=r + 1, min_col=1, max_col=5):
        for cell in row:
            if cell.font == Font():
                cell.font = BODY_FONT


def build_metodologia(wb):
    ws = wb.create_sheet("Metodologia")
    autofit(ws, [95])
    notes = [
        ("Lado oficial (SEFAZ-ES)", (
            "Boletins de 'Distribuição de ICMS/IPVA às Prefeituras Municipais', SEFAZ-ES, Subsecretaria "
            "do Tesouro Estadual (sefaz.es.gov.br/transferencias-constitucionais-aos-municipios), "
            "disponíveis apenas em PDF. Para 2023 e 2024, o valor anual é a soma dos 12 boletins mensais "
            "(sempre a versão republicada/corrigida quando havia mais de uma para o mesmo mês); para "
            "2025, o demonstrativo acumulado anual já consolidado pela própria SEFAZ."
        )),
        ("Lado declarado (SICONFI/SIOPE/SIOPS)", (
            "Compilação, por município do Espírito Santo e ano (2019-2025), do valor de 'Cota-Parte do "
            "ICMS' na DCA (Declaração de Contas Anuais, SICONFI/Tesouro Nacional), no SIOPE (Ministério "
            "da Educação) e no SIOPS (Ministério da Saúde), com sinalização prévia de divergência entre "
            "os três sistemas."
        )),
        ("Correção bruto/líquido", (
            "A comparação inicial entre o valor declarado e o ICMS+IPVA líquidos (após dedução do FUNDEB) "
            "apontava uma diferença agregada sistemática de cerca de 10% ao ano, de magnitude e "
            "uniformidade incompatíveis com erro de preenchimento pontual. A causa identificada foi "
            "conceitual: a Cota-Parte do ICMS é registrada como receita orçamentária pelo valor BRUTO, "
            "antes da dedução do FUNDEB. Restringindo a comparação ao ICMS bruto (sem o IPVA, que a base "
            "declarada não segrega como Cota-Parte do ICMS), o agregado do Espírito Santo reconcilia "
            "quase integralmente nos três anos disponíveis."
        )),
        ("Implicação para o coeficiente da reforma", (
            "Como o agregado estadual reconcilia quase integralmente, o coeficiente de transição do "
            "Espírito Santo como um todo não deveria ser afetado de forma relevante por essas "
            "divergências. O risco identificado é a distribuição interna entre municípios: quando um "
            "ente declara valor sistematicamente distinto do recebido, ou quando os três sistemas "
            "federais divergem entre si, qualquer critério de rateio intramunicipal fundamentado nessas "
            "declarações herda o mesmo erro."
        )),
    ]
    r = 1
    ws.cell(row=r, column=1, value="Notas metodológicas").font = TITLE_FONT
    ws.row_dimensions[r].height = 24
    r += 2
    for title, text in notes:
        ws.cell(row=r, column=1, value=title).font = H_FONT
        r += 1
        cell = ws.cell(row=r, column=1, value=text)
        cell.alignment = WRAP
        cell.font = BODY_FONT
        ws.row_dimensions[r].height = 90
        r += 2


def build_year_sheet(wb, ano):
    ws = wb.create_sheet(str(ano))
    autofit(ws, [30, 20, 20, 16, 14, 16])
    rows = sorted(
        (c for c in DATA["comparacao_municipio_sefaz_siconfi"] if c["ano"] == ano),
        key=lambda c: c["municipio"],
    )
    write_header_row(ws, 1, [
        "Município", "Oficial (SEFAZ), ICMS bruto", "Declarado (SICONFI)",
        "Diferença (R$)", "Diferença (%)", "DCA=SIOPE=SIOPS?",
    ])
    r = 2
    for c in rows:
        ws.cell(row=r, column=1, value=c["municipio"])
        ws.cell(row=r, column=2, value=c["oficial_icms_bruto"]).number_format = RS_FMT
        ws.cell(row=r, column=3, value=c["declarado_siconfi"]).number_format = RS_FMT
        ws.cell(row=r, column=4, value=c["diff_rs"]).number_format = RS_FMT
        ws.cell(row=r, column=5, value=c["diff_pct"] / 100).number_format = PCT_FMT
        ws.cell(row=r, column=6, value=c["validacao_interna"])
        r += 1

    soma_of = sum(c["oficial_icms_bruto"] for c in rows)
    soma_de = sum(c["declarado_siconfi"] for c in rows)
    ws.cell(row=r, column=1, value=f"Total ({len(rows)} municípios)").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=soma_of).number_format = RS_FMT
    ws.cell(row=r, column=3, value=soma_de).number_format = RS_FMT
    ws.cell(row=r, column=4, value=soma_of - soma_de).number_format = RS_FMT
    ws.cell(row=r, column=5, value=(soma_of - soma_de) / soma_de).number_format = PCT_FMT
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = TOTAL_FILL
        ws.cell(row=r, column=col).font = TOTAL_FONT


def build_validacao(wb):
    ws = wb.create_sheet("Validação interna")
    autofit(ws, [30, 10, 18, 18, 18, 14, 12, 16])
    rows = sorted(
        DATA["validacao_interna_dca_siope_siops"],
        key=lambda r: (r["ano"], r["municipio"]),
    )
    write_header_row(ws, 1, [
        "Município", "Ano", "DCA", "SIOPE", "SIOPS", "Validação", "Fonte divergente", "Desvio médio (%)",
    ])
    r = 2
    for row in rows:
        ws.cell(row=r, column=1, value=row["municipio"])
        ws.cell(row=r, column=2, value=row["ano"])
        for col, key in ((3, "dca"), (4, "siope"), (5, "siops")):
            cell = ws.cell(row=r, column=col, value=row[key])
            if row[key] is not None:
                cell.number_format = RS_FMT
        ws.cell(row=r, column=6, value=row["validacao"])
        ws.cell(row=r, column=7, value=row["base_divergencia"])
        ws.cell(row=r, column=8, value=row["desvio_medio_pct"] / 100).number_format = PCT_FMT
        r += 1


def main():
    wb = Workbook()
    build_sintese(wb)
    build_metodologia(wb)
    for ano in (2023, 2024, 2025):
        build_year_sheet(wb, ano)
    build_validacao(wb)
    OUTPUT.parent.mkdir(exist_ok=True)
    wb.save(OUTPUT)
    print(f"Salvo em {OUTPUT} ({OUTPUT.stat().st_size / 1024:.1f} KB)")


if __name__ == "__main__":
    main()
